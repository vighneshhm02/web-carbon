#!/usr/bin/env python3
"""
WebCarbon Research Scanner
===========================
Scans the 70 popular Indian websites listed in websites.txt through
the local WebCarbon API and exports results to a formatted Excel file
suitable for academic paper publication.

Usage:
    python3 scan_websites.py

Requirements:
    - WebCarbon backend running on http://localhost:5000
    - pip install requests openpyxl tqdm
"""

import json
import os
import sys
import time
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
API_BASE        = "http://localhost:5000/api"
SCAN_ENDPOINT   = f"{API_BASE}/scan"
WEBSITES_FILE   = os.path.join(os.path.dirname(__file__), "websites.txt")
DELAY_BETWEEN   = 3          # seconds to wait between each scan
SCAN_TIMEOUT    = 120        # max seconds to wait for a single scan response
OUTPUT_DIR      = os.path.dirname(__file__)

# ─────────────────────────────────────────────
# COLOUR PALETTE (for Excel)
# ─────────────────────────────────────────────
CLR_HEADER_BG   = "1F4E79"   # dark navy
CLR_HEADER_FG   = "FFFFFF"   # white
CLR_SUBHDR_BG   = "2E75B6"   # medium blue
CLR_SUBHDR_FG   = "FFFFFF"
CLR_ROW_EVEN    = "EBF3FB"   # light blue
CLR_ROW_ODD     = "FFFFFF"   # white
CLR_GRADE: dict = {
    "A": "00B050",  # green
    "B": "92D050",  # lime
    "C": "FFFF00",  # yellow
    "D": "FFC000",  # orange
    "E": "FF0000",  # red
    "F": "7030A0",  # purple
}
CLR_FAIL_BG     = "FFE2E2"   # soft red for failed rows
CLR_FAIL_FG     = "CC0000"

THIN_BORDER = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def format_bytes(b):
    """Return a human-readable byte size string."""
    if b is None:
        return "N/A"
    if b >= 1_000_000:
        return f"{b / 1_000_000:.2f} MB"
    if b >= 1_000:
        return f"{b / 1_000:.1f} KB"
    return f"{b} B"


def load_websites(filepath):
    """
    Parse websites.txt.  Returns list of (url, category) tuples.
    Lines starting with # or that are blank are skipped.
    """
    entries = []
    with open(filepath, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split("|")
            url = parts[0].strip()
            category = parts[1].strip() if len(parts) > 1 else "Uncategorized"
            entries.append((url, category))
    return entries


def scan_url(url, timeout=SCAN_TIMEOUT):
    """
    POST to /api/scan. Returns (result_dict | None, error_str | None).
    """
    try:
        resp = requests.post(
            SCAN_ENDPOINT,
            json={"url": url},
            timeout=timeout,
        )
        if resp.status_code == 201:
            return resp.json(), None
        # Non-201 — surface the server error message
        try:
            err = resp.json().get("error", resp.text)
        except Exception:
            err = resp.text
        return None, f"HTTP {resp.status_code}: {err}"
    except requests.Timeout:
        return None, f"Timed out after {timeout}s"
    except requests.ConnectionError:
        return None, "Connection refused — is the backend running on port 5000?"
    except Exception as exc:
        return None, str(exc)


# ─────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────
COLUMNS = [
    # (header label, width)
    ("S.No.",             6),
    ("Website URL",       38),
    ("Category",          20),
    ("Grade",             8),
    ("Score (0–100)",     14),
    ("CO₂/Visit (g)",     14),
    ("CO₂ Source",        14),
    ("Page Size",         14),
    ("Requests",          10),
    ("3rd-Party Scripts", 18),
    ("Image Size",        14),
    ("Font Size",         12),
    ("Response Time (ms)",18),
    ("CDN",               8),
    ("Green Hosting",     14),
    ("Scanned At",        22),
    ("Status",            12),
]

COL_LETTERS = [get_column_letter(i + 1) for i in range(len(COLUMNS))]


def _cell(ws, row, col, value="", bold=False, fg=None, bg=None,
          align="left", wrap=False, border=True, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, color=fg or "000000", size=10)
    if bg:
        c.fill = PatternFill(fill_type="solid", fgColor=bg)
    c.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap
    )
    if border:
        c.border = THIN_BORDER
    if num_fmt:
        c.number_format = num_fmt
    return c


def build_excel(records, timestamp_str):
    wb = Workbook()

    # ── Sheet 1: Raw Data ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Scan Results"

    # Title row
    ws.merge_cells(f"A1:{COL_LETTERS[-1]}1")
    title_cell = ws["A1"]
    title_cell.value = (
        "WebCarbon Research Dataset — CO₂ Footprint of 70 Popular Indian Websites"
    )
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=CLR_HEADER_FG)
    title_cell.fill = PatternFill(fill_type="solid", fgColor=CLR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-header: generated at
    ws.merge_cells(f"A2:{COL_LETTERS[-1]}2")
    sub = ws["A2"]
    sub.value = f"Generated: {timestamp_str}  |  Source: WebCarbon API + Puppeteer CDP Crawl"
    sub.font = Font(name="Calibri", italic=True, size=9, color=CLR_SUBHDR_FG)
    sub.fill = PatternFill(fill_type="solid", fgColor=CLR_SUBHDR_BG)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Column headers (row 3)
    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        c = _cell(
            ws, 3, col_idx, value=label,
            bold=True, fg=CLR_HEADER_FG, bg=CLR_HEADER_BG,
            align="center",
        )
        ws.column_dimensions[COL_LETTERS[col_idx - 1]].width = width
    ws.row_dimensions[3].height = 20

    # Data rows start at row 4
    for row_idx, rec in enumerate(records, start=4):
        bg = CLR_ROW_EVEN if (row_idx % 2 == 0) else CLR_ROW_ODD
        is_fail = rec.get("status") == "Failed"
        if is_fail:
            bg = CLR_FAIL_BG

        def dc(col, value, align="left", bold=False, fg=None, num_fmt=None):
            _cell(ws, row_idx, col, value, bold=bold,
                  fg=fg or (CLR_FAIL_FG if is_fail else "000000"),
                  bg=bg, align=align, num_fmt=num_fmt)

        dc(1,  rec.get("sno"),            align="center")
        dc(2,  rec.get("url"))
        dc(3,  rec.get("category"))

        # Grade cell — colour-coded
        grade = rec.get("grade", "")
        grade_bg  = CLR_GRADE.get(grade, bg) if not is_fail else None
        grade_cell = ws.cell(row=row_idx, column=4, value=grade)
        grade_cell.font = Font(
            name="Calibri", bold=True, size=11,
            color="FFFFFF" if grade in CLR_GRADE else CLR_FAIL_FG
        )
        grade_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=grade_bg if grade_bg else bg
        )
        grade_cell.alignment = Alignment(horizontal="center", vertical="center")
        grade_cell.border = THIN_BORDER

        dc(5,  rec.get("score"),          align="center")
        dc(6,  rec.get("co2_grams"),      align="center",  num_fmt="0.0000")
        dc(7,  rec.get("co2_source"),     align="center")
        dc(8,  rec.get("page_size"),      align="right")
        dc(9,  rec.get("requests"),       align="center")
        dc(10, rec.get("third_party"),    align="center")
        dc(11, rec.get("image_size"),     align="right")
        dc(12, rec.get("font_size"),      align="right")
        dc(13, rec.get("response_time"),  align="center")
        dc(14, rec.get("cdn"),            align="center")
        dc(15, rec.get("green_hosting"),  align="center")
        dc(16, rec.get("scanned_at"))
        dc(17, rec.get("status"),         align="center",
           bold=is_fail, fg=CLR_FAIL_FG if is_fail else "007700")

        ws.row_dimensions[row_idx].height = 16

    # Freeze header rows
    ws.freeze_panes = "A4"

    # ── Sheet 2: Summary Statistics ────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    def sh(row, col, val, bold=False, bg=None, fg=None, align="left"):
        _cell(ws2, row, col, val, bold=bold, bg=bg, fg=fg, align=align)

    ws2.merge_cells("A1:B1")
    ws2["A1"].value = "Summary Statistics"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=13, color=CLR_HEADER_FG)
    ws2["A1"].fill = PatternFill(fill_type="solid", fgColor=CLR_HEADER_BG)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 26

    success_recs = [r for r in records if r.get("status") == "Success"]
    fail_recs    = [r for r in records if r.get("status") == "Failed"]
    co2_vals     = [r["co2_grams"] for r in success_recs if isinstance(r.get("co2_grams"), (int, float))]
    score_vals   = [r["score"]     for r in success_recs if isinstance(r.get("score"),     (int, float))]

    grade_counts = {}
    for r in success_recs:
        g = r.get("grade", "?")
        grade_counts[g] = grade_counts.get(g, 0) + 1

    cat_co2 = {}
    for r in success_recs:
        cat = r.get("category", "Unknown")
        if cat not in cat_co2:
            cat_co2[cat] = []
        if isinstance(r.get("co2_grams"), (int, float)):
            cat_co2[cat].append(r["co2_grams"])

    stats = [
        ("── SCAN OVERVIEW ──",          "",                                True,  CLR_SUBHDR_BG, CLR_SUBHDR_FG),
        ("Total Websites",               len(records),                      False, None, None),
        ("Successfully Scanned",         len(success_recs),                 False, None, None),
        ("Failed Scans",                 len(fail_recs),                    False, None, None),
        ("── CO₂ STATISTICS (g/visit) ──","",                              True,  CLR_SUBHDR_BG, CLR_SUBHDR_FG),
        ("Minimum CO₂",                  f"{min(co2_vals):.4f}" if co2_vals else "N/A", False, None, None),
        ("Maximum CO₂",                  f"{max(co2_vals):.4f}" if co2_vals else "N/A", False, None, None),
        ("Mean CO₂",                     f"{sum(co2_vals)/len(co2_vals):.4f}" if co2_vals else "N/A", False, None, None),
        ("Median CO₂",                   f"{sorted(co2_vals)[len(co2_vals)//2]:.4f}" if co2_vals else "N/A", False, None, None),
        ("── SUSTAINABILITY SCORE ──",   "",                                True,  CLR_SUBHDR_BG, CLR_SUBHDR_FG),
        ("Min Score",                    f"{min(score_vals):.1f}" if score_vals else "N/A", False, None, None),
        ("Max Score",                    f"{max(score_vals):.1f}" if score_vals else "N/A", False, None, None),
        ("Mean Score",                   f"{sum(score_vals)/len(score_vals):.1f}" if score_vals else "N/A", False, None, None),
        ("── GRADE DISTRIBUTION ──",     "",                                True,  CLR_SUBHDR_BG, CLR_SUBHDR_FG),
    ]
    for g in ["A", "B", "C", "D", "E", "F"]:
        stats.append((f"Grade {g}", grade_counts.get(g, 0), False, None, None))

    stats.append(("── CO₂ BY CATEGORY ──",      "",   True, CLR_SUBHDR_BG, CLR_SUBHDR_FG))
    for cat, vals in sorted(cat_co2.items()):
        avg = sum(vals) / len(vals) if vals else 0
        stats.append((cat, f"{avg:.4f} g avg ({len(vals)} sites)", False, None, None))

    for i, (label, value, bold, bg, fg) in enumerate(stats, start=2):
        sh(i, 1, label, bold=bold, bg=bg or (CLR_ROW_EVEN if i % 2 == 0 else CLR_ROW_ODD), fg=fg)
        sh(i, 2, value, bold=bold, bg=bg or (CLR_ROW_EVEN if i % 2 == 0 else CLR_ROW_ODD), fg=fg, align="center")
        ws2.row_dimensions[i].height = 16

    # ── Sheet 3: Category Breakdown ───────────────────────────────────
    ws3 = wb.create_sheet("By Category")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 38

    ws3.merge_cells("A1:C1")
    ws3["A1"].value = "Category-wise CO₂ Rankings"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=13, color=CLR_HEADER_FG)
    ws3["A1"].fill = PatternFill(fill_type="solid", fgColor=CLR_HEADER_BG)
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 26

    for col, label in enumerate(["Category", "CO₂ (g/visit)", "Website URL"], start=1):
        _cell(ws3, 2, col, label, bold=True, fg=CLR_HEADER_FG, bg=CLR_SUBHDR_BG, align="center")
    ws3.row_dimensions[2].height = 18

    sorted_by_co2 = sorted(
        success_recs,
        key=lambda r: r.get("co2_grams") if isinstance(r.get("co2_grams"), (int, float)) else 999
    )
    for i, r in enumerate(sorted_by_co2, start=3):
        bg = CLR_ROW_EVEN if i % 2 == 0 else CLR_ROW_ODD
        _cell(ws3, i, 1, r.get("category", ""), bg=bg)
        _cell(ws3, i, 2, r.get("co2_grams", ""), align="center", bg=bg, num_fmt="0.0000")
        _cell(ws3, i, 3, r.get("url", ""), bg=bg)
        ws3.row_dimensions[i].height = 15

    ws3.freeze_panes = "A3"

    return wb


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print("\n╔══════════════════════════════════════════════════╗")
    print("║   WebCarbon — Batch Website Scanner              ║")
    print("║   For Academic Research / Paper Publication      ║")
    print("╚══════════════════════════════════════════════════╝\n")

    # Load website list
    if not os.path.exists(WEBSITES_FILE):
        print(f"[ERROR] websites.txt not found at: {WEBSITES_FILE}")
        sys.exit(1)

    websites = load_websites(WEBSITES_FILE)
    print(f"[INFO] Loaded {len(websites)} websites from websites.txt")

    # Check backend is up
    print("[INFO] Checking backend connectivity...")
    try:
        resp = requests.get(f"{API_BASE}/history", timeout=5)
        print(f"[INFO] Backend is reachable (HTTP {resp.status_code})")
    except Exception as e:
        print(f"[ERROR] Cannot reach backend at {API_BASE}: {e}")
        print("       Start the backend with:  cd server && npm run dev")
        sys.exit(1)

    # Timestamp for file naming
    ts = datetime.now()
    timestamp_str = ts.strftime("%Y-%m-%d %H:%M:%S")
    file_ts       = ts.strftime("%Y%m%d_%H%M%S")
    output_file   = os.path.join(OUTPUT_DIR, f"webcarbon_india_{file_ts}.xlsx")

    print(f"\n[INFO] Starting scan of {len(websites)} websites...")
    print(f"[INFO] Estimated time: ~{len(websites) * (15 + DELAY_BETWEEN) // 60}–{len(websites) * (30 + DELAY_BETWEEN) // 60} minutes\n")

    records = []
    failed_urls = []

    with tqdm(total=len(websites), unit="site", colour="green") as pbar:
        for sno, (url, category) in enumerate(websites, start=1):
            pbar.set_description(f"Scanning {url[:45]}")

            result, error = scan_url(url)

            if result:
                m = result.get("metrics", {})
                c = result.get("co2", {})
                records.append({
                    "sno":           sno,
                    "url":           url,
                    "category":      category,
                    "grade":         result.get("grade", ""),
                    "score":         result.get("score"),
                    "co2_grams":     c.get("grams"),
                    "co2_source":    c.get("source", "").upper(),
                    "page_size":     format_bytes(m.get("totalBytes")),
                    "requests":      m.get("totalRequests"),
                    "third_party":   m.get("thirdPartyScripts"),
                    "image_size":    format_bytes(m.get("imageSize")),
                    "font_size":     format_bytes(m.get("fontSize")),
                    "response_time": m.get("responseTime"),
                    "cdn":           "Yes" if m.get("isCDN") else "No",
                    "green_hosting": "Yes" if m.get("isGreenHosting") else "No",
                    "scanned_at":    result.get("scannedAt", timestamp_str),
                    "status":        "Success",
                })
                pbar.set_postfix(grade=result.get("grade"), co2=f"{c.get('grams', '?')}g")
            else:
                records.append({
                    "sno":      sno,
                    "url":      url,
                    "category": category,
                    "grade":    "",
                    "score":    "",
                    "co2_grams":     "",
                    "co2_source":    "",
                    "page_size":     "",
                    "requests":      "",
                    "third_party":   "",
                    "image_size":    "",
                    "font_size":     "",
                    "response_time": "",
                    "cdn":           "",
                    "green_hosting": "",
                    "scanned_at":    timestamp_str,
                    "status":        "Failed",
                })
                failed_urls.append((url, error))
                pbar.set_postfix(status="FAILED")

            pbar.update(1)

            # Delay between scans to avoid hammering Puppeteer
            if sno < len(websites):
                time.sleep(DELAY_BETWEEN)

    # Build and save Excel
    print("\n[INFO] Building Excel report...")
    wb = build_excel(records, timestamp_str)
    wb.save(output_file)

    # Print summary
    success = sum(1 for r in records if r["status"] == "Success")
    fail    = len(records) - success
    print(f"\n{'─'*52}")
    print(f"  ✓ Successfully scanned : {success}/{len(records)}")
    print(f"  ✗ Failed               : {fail}/{len(records)}")
    print(f"{'─'*52}")

    if failed_urls:
        print("\n  Failed URLs:")
        for url, err in failed_urls:
            print(f"    • {url}")
            print(f"      └─ {err}")

    print(f"\n  📊 Excel saved to:")
    print(f"     {output_file}\n")

    co2_vals = [r["co2_grams"] for r in records if isinstance(r.get("co2_grams"), (int, float))]
    if co2_vals:
        print(f"  CO₂ range  : {min(co2_vals):.4f}g – {max(co2_vals):.4f}g per visit")
        print(f"  Mean CO₂   : {sum(co2_vals)/len(co2_vals):.4f}g per visit")
    print()


if __name__ == "__main__":
    main()
